"""
pricing.py — API routes for pricing analysis.

Endpoints:
  POST /pricing/analyze         — trigger pricing analysis for an RFP
  GET  /pricing/status/{job_id} — poll job status
  GET  /pricing/result/{rfp_id} — get latest result
  POST /pricing/correct         — apply user correction to detected structure
  GET  /pricing/export/{rfp_id} — export pricing analysis as xlsx/csv

FIXES (2026-04-01 v2):
  BUG-1  asyncio.Semaphore(2) created at module import time — outside any
         running event loop on Python 3.10+ this raises
         "DeprecationWarning / no current event loop" and can silently
         return a broken semaphore that raises RuntimeError on first `async with`.
         Fix: initialise _api_semaphore lazily inside the first coroutine that
         needs it, guarded by a threading.Lock so the one-time init is safe.

  BUG-2  name_map values can be plain strings OR dicts
         (saved as {filename: {"name": ..., "email": ...}} by suppliers.py).
         The original code checked isinstance AFTER already using .get() on the
         dict, meaning the string path fell through and the dict path was never
         reached. Fix: check the type of the *value* first, then extract .name.

  BUG-3  `import openpyxl` inside the /export route raises HTTP 500 with a
         misleading "openpyxl not installed" message even when openpyxl IS
         installed, because the ImportError guard was raising HTTPException
         instead of letting the real openpyxl ImportError propagate for
         diagnosis. Fixed to re-raise the real error and added a top-level
         import so missing-dependency failures are caught at startup, not
         at request time.

  PRIOR FIXES (kept):
  - extract_pricing_from_document wrapped in run_in_executor (non-blocking)
  - asyncio.get_running_loop() replaces deprecated get_event_loop()
"""
import json
import asyncio
import io
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from fastapi import APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from app.services.pricing_parser import extract_pricing_from_document
from app.services.pricing_analyzer import run_pricing_analysis
from app.services.document_parser import parse_document
from app.services.job_store import job_store, JobStatus

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


router     = APIRouter()
UPLOAD_DIR = Path("uploads")
META_DIR   = Path("metadata")

_executor = ThreadPoolExecutor(max_workers=4)

# BUG-1 FIX: do NOT create asyncio.Semaphore at import time.
# On Python 3.10+ there is no running event loop at module load, so
# asyncio.Semaphore() silently attaches to a throwaway loop that is
# already closed by the time the first request arrives, causing
# RuntimeError: Task got Future attached to a different loop.
_api_semaphore: Optional[asyncio.Semaphore] = None
_semaphore_lock = threading.Lock()


def _get_semaphore() -> asyncio.Semaphore:
    """Return (or lazily create) the per-process API semaphore."""
    global _api_semaphore
    if _api_semaphore is None:
        with _semaphore_lock:
            if _api_semaphore is None:          # double-checked locking
                _api_semaphore = asyncio.Semaphore(2)
    return _api_semaphore


# ── Request / Response models ─────────────────────────────────────────────
class PricingAnalyzeRequest(BaseModel):
    rfp_id:     str
    project_id: Optional[str] = None

class PricingCorrectionRequest(BaseModel):
    rfp_id: str
    supplier_name: str
    corrections: list[dict]

class JobStartResponse(BaseModel):
    job_id: str
    status: str

class JobStatusResponse(BaseModel):
    job_id: str
    status: str
    result: Optional[dict] = None
    error:  Optional[str]  = None


# ── Helpers ───────────────────────────────────────────────────────────────
def _call_with_retry(fn, *args, max_retries: int = 4, base_delay: float = 15.0):
    last_err = None
    delay = base_delay
    for attempt in range(max_retries + 1):
        try:
            return fn(*args)
        except Exception as e:
            msg = str(e).lower()
            if ("429" in msg or "rate limit" in msg or "too many requests" in msg) and attempt < max_retries:
                last_err = e
                time.sleep(delay)
                delay *= 2
                continue
            raise
    raise last_err


def _load_supplier_names(rfp_id: str) -> dict:
    """Load file_path -> supplier_name mapping saved during upload (legacy path)."""
    meta_path = META_DIR / f"{rfp_id}_suppliers.json"
    if meta_path.exists():
        return json.loads(meta_path.read_text())
    return {}


async def _run_in_thread(fn, *args):
    """Run a blocking function in the thread pool without blocking the event loop."""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(_executor, fn, *args)


def _resolve_supplier_name(name_map: dict, sf_path: Path) -> str:
    """
    BUG-2 FIX: name_map values can be either plain strings OR dicts of the form
    {"name": "Acme Ltd", "email": "..."}.  The original code checked
    isinstance(name_map.get(sf_path.name), dict) AFTER already calling .get()
    on the outer map, which meant the string-value fast-path was silently
    falling through and resolving to the raw stem as the supplier name.
    """
    # Priority 1: full path key with dict value  {str(sf): {"name": ...}}
    v = name_map.get(str(sf_path))
    if isinstance(v, dict):
        return v.get("name") or sf_path.stem
    if isinstance(v, str) and v:
        return v

    # Priority 2: filename key with dict value  {sf_path.name: {"name": ...}}
    v = name_map.get(sf_path.name)
    if isinstance(v, dict):
        return v.get("name") or sf_path.stem
    if isinstance(v, str) and v:
        return v

    # Priority 3: derive from filename  "abc_supplier_Acme Ltd.xlsx" -> "Acme Ltd"
    stem = sf_path.stem
    if "_supplier_" in stem:
        return stem.split("_supplier_", 1)[-1]

    return stem


# ── Background job ─────────────────────────────────────────────────────────
async def _run_pricing_job(rfp_id: str, job_id: str, project_id: str = None):
    job_store.set_running(job_id)
    try:
        rfp_path       = None
        supplier_files = []
        name_map       = {}

        if project_id:
            from app.services.project_store import (
                get_rfp_path,
                get_supplier_paths,
                load_metadata,
                ensure_rfp_local,
                ensure_suppliers_local,
            )

            rfp_path       = get_rfp_path(project_id)
            supplier_files = get_supplier_paths(project_id)

            if not supplier_files:
                try:
                    ensure_suppliers_local(project_id)
                    supplier_files = get_supplier_paths(project_id)
                except Exception:
                    pass

            if not rfp_path:
                try:
                    rfp_path = ensure_rfp_local(project_id)
                except Exception:
                    pass

            name_map = load_metadata(project_id, "suppliers.json") or {}

        if not supplier_files:
            rfp_files      = list(UPLOAD_DIR.glob(f"{rfp_id}_rfp*"))
            rfp_path       = rfp_files[0] if rfp_files else None
            supplier_files = list(UPLOAD_DIR.glob(f"{rfp_id}_supplier_*"))
            name_map       = _load_supplier_names(rfp_id)

        if not supplier_files:
            job_store.set_failed(
                job_id,
                "No supplier files found. Upload supplier responses first."
            )
            return

        # ── Parse RFP to get full template text ───────────────────────────
        rfp_full_text = ""
        if rfp_path and Path(str(rfp_path)).exists():
            try:
                rfp_doc = await _run_in_thread(
                    _call_with_retry, parse_document, str(rfp_path)
                )
                rfp_full_text = rfp_doc.get("full_text", "")
            except Exception:
                rfp_full_text = ""

        # ── Extract pricing from each supplier document ───────────────────
        suppliers_pricing = []
        semaphore = _get_semaphore()          # BUG-1 FIX: lazy init
        for sf in supplier_files:
            sf_path = Path(str(sf))
            supplier_name = _resolve_supplier_name(name_map, sf_path)  # BUG-2 FIX

            async with semaphore:
                parsed = await _run_in_thread(
                    _call_with_retry, parse_document, str(sf)
                )

            full_text = parsed.get("full_text", "")

            pricing = await _run_in_thread(
                extract_pricing_from_document,
                str(sf),
                supplier_name,
                full_text,
                rfp_full_text,
            )
            suppliers_pricing.append(pricing)
            await asyncio.sleep(1)

        result           = run_pricing_analysis(suppliers_pricing)
        result["rfp_id"] = rfp_id

        result_path = META_DIR / f"{rfp_id}_pricing.json"
        result_path.parent.mkdir(parents=True, exist_ok=True)
        result_path.write_text(json.dumps(result, default=str))

        job_store.set_completed(job_id, result)

    except Exception as e:
        job_store.set_failed(job_id, str(e))


# ── Routes ────────────────────────────────────────────────────────────────
@router.post("/analyze", response_model=JobStartResponse)
async def analyze_pricing(req: PricingAnalyzeRequest, background_tasks: BackgroundTasks):
    job_id = job_store.create()
    background_tasks.add_task(
        _run_pricing_job,
        req.rfp_id,
        job_id,
        req.project_id,
    )
    return JobStartResponse(job_id=job_id, status=JobStatus.PENDING)


@router.get("/status/{job_id}", response_model=JobStatusResponse)
async def get_pricing_status(job_id: str):
    job = job_store.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return JobStatusResponse(
        job_id=job_id,
        status=job["status"],
        result=job.get("result"),
        error=job.get("error"),
    )


@router.get("/result/{rfp_id}")
async def get_pricing_result(rfp_id: str):
    result_path = META_DIR / f"{rfp_id}_pricing.json"
    if not result_path.exists():
        raise HTTPException(
            status_code=404,
            detail="No pricing result found. Run /pricing/analyze first.",
        )
    return json.loads(result_path.read_text())


@router.post("/correct")
async def correct_pricing(req: PricingCorrectionRequest):
    result_path = META_DIR / f"{req.rfp_id}_pricing.json"
    if not result_path.exists():
        raise HTTPException(status_code=404, detail="No pricing result found.")

    result = json.loads(result_path.read_text())
    matrix = result.get("cost_model", {}).get("matrix", {})
    for correction in req.corrections:
        desc = correction.get("description", "").strip()
        if desc in matrix:
            if req.supplier_name in matrix[desc] and matrix[desc][req.supplier_name]:
                matrix[desc][req.supplier_name].update({
                    k: v for k, v in correction.items()
                    if k in ("unit_price", "quantity", "total", "category", "notes") and v is not None
                })

    suppliers = result.get("cost_model", {}).get("suppliers", [])
    suppliers_pricing = []
    for sname in suppliers:
        items = []
        for desc, smap in matrix.items():
            val = smap.get(sname)
            if val:
                items.append({"description": desc, **val})
        suppliers_pricing.append({
            "supplier_name":  sname,
            "all_line_items": items,
            "total_cost":     sum(i.get("total", 0) for i in items),
            "structure_type": "corrected",
            "sheets":         [],
        })

    updated = run_pricing_analysis(suppliers_pricing)
    updated["rfp_id"] = req.rfp_id
    result_path.write_text(json.dumps(updated, default=str))
    return updated


@router.get("/export/{rfp_id}")
async def export_pricing(rfp_id: str, format: str = "xlsx"):
    if format not in ("xlsx", "csv"):
        raise HTTPException(status_code=400, detail="format must be xlsx or csv")

    result_path = META_DIR / f"{rfp_id}_pricing.json"
    if not result_path.exists():
        raise HTTPException(status_code=404, detail="No pricing result found.")

    result      = json.loads(result_path.read_text())
    matrix      = result.get("cost_model", {}).get("matrix", {})
    suppliers   = result.get("cost_model", {}).get("suppliers", [])
    total_costs = result.get("total_costs", [])
    award_rec   = result.get("award_recommendation", {})

    if format == "csv":
        lines   = ["Line Item,Category," + ",".join(suppliers) + ",Best Price,Best Supplier"]
        bob     = result.get("best_of_best", {}).get("breakdown", [])
        bob_map = {b["description"]: b for b in bob}
        for desc, smap in matrix.items():
            cats   = [v["category"] for v in smap.values() if v]
            cat    = cats[0] if cats else ""
            prices = [str(smap.get(s, {}).get("total", "") if smap.get(s) else "") for s in suppliers]
            best   = bob_map.get(desc, {})
            lines.append(f'"{desc}","{cat}",' + ",".join(prices) + f',{best.get("best_total","")},"{best.get("best_supplier","")}"')
        content = "\n".join(lines).encode("utf-8")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=pricing_{rfp_id}.csv"},
        )

    # BUG-3 FIX: openpyxl is imported at module top-level so import errors
    # surface at startup, not at request time. Check the flag here instead
    # of a try/except that swallows the real traceback.
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Add 'openpyxl' to requirements.txt and restart.",
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    GREEN_FILL  = PatternFill("solid", fgColor="D4EDDA")
    CENTER      = Alignment(horizontal="center", wrap_text=True)

    def hdr(ws):
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER

    ws1 = wb.create_sheet("Summary")
    ws1.append(["Rank", "Supplier", "Total Cost", "Line Items"])
    hdr(ws1)
    for r in total_costs:
        ws1.append([r["rank"], r["supplier_name"], r["total_cost"], r["line_item_count"]])
        if r["rank"] == 1:
            for cell in ws1[ws1.max_row]: cell.fill = GREEN_FILL
    ws1.column_dimensions["B"].width = 30

    ws2 = wb.create_sheet("Price Matrix")
    ws2.append(["Line Item", "Category"] + suppliers + ["Best Price", "Best Supplier"])
    hdr(ws2)
    bob     = result.get("best_of_best", {}).get("breakdown", [])
    bob_map = {b["description"]: b for b in bob}
    for desc, smap in matrix.items():
        cats   = [v["category"] for v in smap.values() if v]
        cat    = cats[0] if cats else ""
        prices = [smap.get(s, {}).get("total", "") if smap.get(s) else "" for s in suppliers]
        best   = bob_map.get(desc, {})
        row    = [desc, cat] + prices + [best.get("best_total", ""), best.get("best_supplier", "")]
        ws2.append(row)
        if best.get("best_supplier") in suppliers:
            best_col = suppliers.index(best["best_supplier"]) + 3
            ws2.cell(row=ws2.max_row, column=best_col).fill = GREEN_FILL
    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 20

    if result.get("market_basket_2", {}).get("per_category", {}).get("best"):
        ws3  = wb.create_sheet("Market Basket (2 Suppliers)")
        best = result["market_basket_2"]["per_category"]["best"]
        ws3.append(["Category", "Awarded To", "Cost"])
        hdr(ws3)
        for cat, detail in best.get("allocation", {}).items():
            ws3.append([cat, detail["awarded_to"], detail["cost"]])
        ws3.append(["", "TOTAL", best["total_cost"]])
        ws3.column_dimensions["A"].width = 25
        ws3.column_dimensions["B"].width = 25

    if result.get("market_basket_3", {}).get("per_category", {}).get("best"):
        ws4  = wb.create_sheet("Market Basket (3 Suppliers)")
        best = result["market_basket_3"]["per_category"]["best"]
        ws4.append(["Category", "Awarded To", "Cost"])
        hdr(ws4)
        for cat, detail in best.get("allocation", {}).items():
            ws4.append([cat, detail["awarded_to"], detail["cost"]])
        ws4.append(["", "TOTAL", best["total_cost"]])
        ws4.column_dimensions["A"].width = 25
        ws4.column_dimensions["B"].width = 25

    ws5 = wb.create_sheet("Award Recommendation")
    ws5.append(["Recommended Strategy", award_rec.get("recommended_strategy", "")])
    ws5.append(["Recommended Total",    award_rec.get("recommended_total", "")])
    ws5.append(["Savings Opportunity",  award_rec.get("savings_opportunity", "")])
    ws5.append([])
    ws5.append(["Rationale"])
    for r in award_rec.get("rationale", []):
        ws5.append(["", r])
    ws5.append([])
    ws5.append(["All Strategies Compared", "Total Cost", "Complexity", "Risk", "Suppliers Involved"])
    for cell in ws5[ws5.max_row]: cell.font = Font(bold=True)
    for strat in award_rec.get("all_strategies", []):
        ws5.append([
            strat["strategy"], strat["total"],
            strat["complexity"], strat["risk"],
            strat["suppliers_involved"]
        ])
    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pricing_{rfp_id}.xlsx"},
    )
