"""
technical_parser.py — Robust parser for technical question files (FORMAT A & B).

Handles two complex embedded layouts:
  FORMAT A: Multi-section flat sheet (one sheet per supplier)
  FORMAT B: Multi-sheet single supplier (sheet1=pricing, sheet2=technical, sheet3=supplier info)

Returns parsed questions with:
  - Supplier name (inferred from header, SI-01, response patterns, or filename)
  - Score hints (from compliance status: 0.0–1.0)
  - Response quality (full|template|empty)
  - Proper section detection and pricing section skipping
"""

import re
import logging
from io import BytesIO
from pathlib import Path
from typing import Optional, Dict, List, Any

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    openpyxl = None
    Worksheet = None
    InvalidFileException = Exception

logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════════════════════════
# FORMAT DETECTION
# ════════════════════════════════════════════════════════════════════════════════

def detect_format(wb: "openpyxl.Workbook") -> str:
    """
    Detect file format by looking for RFP RESPONSE or Supplier: header in rows 1-3.
    Returns "A" if found (FORMAT A), else "B" (FORMAT B).
    """
    for ws in wb.worksheets:
        for row_idx in range(1, 4):
            for cell in ws[row_idx]:
                if cell.value is None:
                    continue
                val = str(cell.value).strip().lower()
                if "rfp response" in val or val.startswith("supplier:"):
                    return "A"
    return "B"


# ════════════════════════════════════════════════════════════════════════════════
# SECTION DETECTION
# ════════════════════════════════════════════════════════════════════════════════

def detect_sections(ws: "Worksheet") -> List[Dict[str, Any]]:
    """
    Find sections in a worksheet.

    Returns: [{"section_name": str, "header_row_idx": int, "data_start_idx": int, "data_end_idx": int}, ...]

    A section header is a row where:
      - First non-empty cell contains "SECTION" (case-insensitive) AND
      - Row has < 3 non-empty cells

    For FORMAT B (no section headers): treat entire sheet as one section.
    """
    sections = []
    max_row = ws.max_row or 1

    # Check if this is a FORMAT B sheet (no section headers, starts with Q# or SKU#)
    has_section_headers = False
    first_non_empty_row = None

    for row_idx in range(1, min(max_row + 1, 20)):  # Check first 20 rows
        non_empty = [c.value for c in ws[row_idx] if c.value is not None]
        if not non_empty:
            continue
        first_non_empty_row = row_idx
        first_val = str(non_empty[0]).strip().lower()
        if "section" in first_val and len(non_empty) < 3:
            has_section_headers = True
            break

    if not has_section_headers and first_non_empty_row:
        # FORMAT B: no section headers, treat entire sheet as one section
        # Find the Q# header row
        header_row_idx = None
        for row_idx in range(first_non_empty_row, min(max_row + 1, first_non_empty_row + 10)):
            cell_val = str(ws.cell(row_idx, 1).value or "").strip()
            if cell_val.lower() in ("q#", "q #", "question_id", "qid") or cell_val.upper() == "Q#":
                header_row_idx = row_idx
                break

        if header_row_idx:
            # Skip if first cell of row 1 is "SKU#" (pricing sheet)
            if str(ws.cell(1, 1).value or "").strip().upper() == "SKU#":
                return []

            return [{
                "section_name": ws.title,
                "header_row_idx": header_row_idx,
                "data_start_idx": header_row_idx + 1,
                "data_end_idx": max_row,
            }]
        return []

    # FORMAT A: Has section headers
    current_row = 1
    while current_row <= max_row:
        # Find next section header
        header_row_idx = None
        section_name = None

        for row_idx in range(current_row, max_row + 1):
            non_empty = [c for c in ws[row_idx] if c.value is not None]
            if not non_empty:
                continue

            first_val = str(non_empty[0]).strip().lower()
            non_empty_count = len(non_empty)

            # Section header: contains "SECTION" and has < 3 non-empty cells
            if "section" in first_val and non_empty_count < 3:
                # Skip pricing sections
                if any(skip in first_val for skip in ["pricing", "commercial", "sku", "section c"]):
                    current_row = row_idx + 1
                    continue

                header_row_idx = row_idx
                section_name = str(non_empty[0]).strip()
                break

        if not header_row_idx:
            break

        # Find column headers row (first row after section header with >= 3 non-empty cells)
        col_header_idx = None
        for row_idx in range(header_row_idx + 1, max_row + 1):
            non_empty = [c for c in ws[row_idx] if c.value is not None]
            if len(non_empty) >= 3:
                col_header_idx = row_idx
                break

        if not col_header_idx:
            current_row = header_row_idx + 1
            continue

        # Find data end (next section header or end of sheet)
        data_end_idx = max_row
        for row_idx in range(col_header_idx + 1, max_row + 1):
            non_empty = [c for c in ws[row_idx] if c.value is not None]
            if not non_empty:
                continue

            first_val = str(non_empty[0]).strip().lower()
            if "section" in first_val and len(non_empty) < 3:
                data_end_idx = row_idx - 1
                break

        sections.append({
            "section_name": section_name,
            "header_row_idx": header_row_idx,
            "data_start_idx": col_header_idx + 1,
            "data_end_idx": data_end_idx,
        })

        current_row = data_end_idx + 1

    return sections


# ════════════════════════════════════════════════════════════════════════════════
# COLUMN MAPPING
# ════════════════════════════════════════════════════════════════════════════════

def map_columns(header_row: List[Any]) -> Dict[str, int]:
    """
    Map column index to canonical field names.

    Returns: {"question_id": col_idx, "category": col_idx, "question_text": col_idx, ...}

    Canonical names and their patterns:
      - "question_id": "Q#", "Q #", "question_id", "qid"
      - "category": "Category", "Section", "Topic"
      - "question_text": "Question", "Question Text", "Requirement"
      - "response": "Supplier Response", "Response", "Answer", "Reply"
      - "compliance": "Compliance", "Completeness", "Status", "Compliance Status"
      - "doc_ref": "Doc Ref", "Supporting Doc Ref", "Document Reference", "Reference"
    """
    mapping = {}

    patterns = {
        "question_id": [r"^q\s*#?$", r"^q\s+#$", r"^question_id$", r"^qid$"],
        "category": [r"^category$", r"^section$", r"^topic$"],
        "question_text": [r"^question", r"^question\s+text$", r"^requirement$"],
        "response": [r"^supplier\s+response$", r"^response$", r"^answer$", r"^reply$"],
        "compliance": [r"^compliance", r"^completeness$", r"^status", r"^compliance\s+status$"],
        "doc_ref": [r"^doc\s+ref", r"^supporting\s+doc", r"^document\s+reference", r"^reference$"],
    }

    for col_idx, cell in enumerate(header_row):
        if cell is None:
            continue

        cell_str = str(cell).strip().lower()

        for canonical, patterns_list in patterns.items():
            for pattern in patterns_list:
                if re.match(pattern, cell_str, re.IGNORECASE):
                    if canonical not in mapping:
                        mapping[canonical] = col_idx
                    break

    return mapping


# ════════════════════════════════════════════════════════════════════════════════
# SUPPLIER NAME EXTRACTION
# ════════════════════════════════════════════════════════════════════════════════

def extract_supplier_name(ws: "Worksheet", sheet_name: str, filename: str) -> str:
    """
    Extract supplier name via priority order:
      1. Scan rows 1-5 for "Supplier: {NAME}" pattern
      2. Look for SI-01 row (legal company name)
      3. Scan response column for "per {NAME} QMS" patterns
      4. Use sheet_name if not generic "Sheet{N}"
      5. Fallback: filename without extension
    """
    # Priority 1: Scan rows 1-5 for "Supplier: {NAME}" pattern
    for row_idx in range(1, 6):
        for cell in ws[row_idx]:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            match = re.search(r"Supplier\s*:\s*(.+?)(?:\s*\||\s*$)", val, re.IGNORECASE)
            if match:
                supplier = match.group(1).strip()
                if supplier and len(supplier) > 1:
                    return supplier

    # Priority 2: Look for SI-01 row
    col_mapping = {}
    for row_idx in range(1, min(ws.max_row + 1, 50)):
        for col_idx, cell in enumerate(ws[row_idx]):
            if cell.value and re.match(r"^si\s*-?\s*01$", str(cell.value).strip(), re.IGNORECASE):
                # Found SI-01 column, get the response from a nearby row
                for check_row in range(row_idx, min(row_idx + 5, ws.max_row + 1)):
                    resp_cell = ws.cell(check_row, col_idx + 3)  # Assume response is ~3 cols right
                    if resp_cell.value:
                        supplier = str(resp_cell.value).strip()
                        if supplier and len(supplier) > 1:
                            return supplier

    # Priority 3: Scan for "per {NAME} QMS" pattern in responses
    name_freq = {}
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            text = str(cell).strip()
            matches = re.findall(r"per\s+([A-Za-z][A-Za-z0-9\s]{2,30}?)(?:\s+QMS|\s+standard|\s+quality)", text, re.IGNORECASE)
            for match in matches:
                match = match.strip()
                if match not in name_freq:
                    name_freq[match] = 0
                name_freq[match] += 1

    if name_freq:
        most_common = max(name_freq.items(), key=lambda x: x[1])
        return most_common[0]

    # Priority 4: Use sheet_name if not generic
    if not re.match(r"^sheet\d+$", sheet_name, re.IGNORECASE):
        return sheet_name.strip()

    # Priority 5: Fallback to filename
    return Path(filename).stem.strip()


# ════════════════════════════════════════════════════════════════════════════════
# COMPLIANCE TO SCORE HINT
# ════════════════════════════════════════════════════════════════════════════════

def compliance_to_score_hint(compliance_value: Any) -> Dict[str, Any]:
    """
    Convert compliance/completeness cell value to score hint (0.0–1.0) and status.

    Returns: {"score_hint": float, "status": "pass" | "partial" | "fail" | "unknown"}
    """
    if compliance_value is None or (isinstance(compliance_value, str) and not compliance_value.strip()):
        return {"score_hint": 0.3, "status": "unknown"}

    val = str(compliance_value).strip().lower()

    # Non-compliant (check FIRST, before "compliant")
    if any(x in val for x in ["non-compliant", "not compliant", "✗"]):
        return {"score_hint": 0.0, "status": "fail"}

    # Full compliance
    if any(x in val for x in ["fully compliant", "✓", "complete"]):
        return {"score_hint": 1.0, "status": "pass"}

    # Handle simple "compliant" (must come after non-compliant check)
    if "compliant" in val:
        return {"score_hint": 1.0, "status": "pass"}

    # Partial compliance
    if any(x in val for x in ["⚠", "partial"]):
        return {"score_hint": 0.5, "status": "partial"}

    # Not applicable (treat as partial, not fail)
    if "not applicable" in val or "n/a" in val:
        return {"score_hint": 0.7, "status": "partial"}

    # Default: partial for any other non-empty value
    if val:
        return {"score_hint": 0.6, "status": "partial"}

    return {"score_hint": 0.3, "status": "unknown"}


# ════════════════════════════════════════════════════════════════════════════════
# RESPONSE QUALITY ASSESSMENT
# ════════════════════════════════════════════════════════════════════════════════

def assess_response_quality(response_text: Any) -> Dict[str, Any]:
    """
    Assess response quality without LLM.

    Returns: {"quality": "full" | "template" | "empty", "char_count": int}
    """
    if response_text is None:
        return {"quality": "empty", "char_count": 0}

    text = str(response_text).strip()
    char_count = len(text)

    if not text:
        return {"quality": "empty", "char_count": 0}

    # Template response patterns
    # Pattern 1: [Company standard response for ...] or [... standard response for ...]
    # Pattern 2: Compliant per ..., Full response: see ...
    template_patterns = [
        r"\[.*?standard.*?response.*?\]",
        r"Compliant\s+per\s+.+\.\s+Full\s+response",
    ]

    if any(re.search(p, text, re.IGNORECASE) for p in template_patterns):
        return {"quality": "template", "char_count": char_count}

    # Full response (must be at least 10 chars)
    if char_count >= 10:
        return {"quality": "full", "char_count": char_count}

    # Short response (less than 10 chars)
    return {"quality": "empty", "char_count": char_count}


# ════════════════════════════════════════════════════════════════════════════════
# CATEGORY INFERENCE
# ════════════════════════════════════════════════════════════════════════════════

def infer_category_from_qid(qid: str) -> str:
    """Infer category from question ID prefix."""
    if not qid:
        return "General"

    prefix = qid[:3].upper()

    category_map = {
        "SI-": "Supplier Information",
        "TQ-": "Technical Questions",
        "SC-": "Supply Chain",
        "QM-": "Quality Management",
    }

    for key, cat in category_map.items():
        if qid.upper().startswith(key):
            return cat

    return "General"


# ════════════════════════════════════════════════════════════════════════════════
# MAIN PARSE FUNCTION
# ════════════════════════════════════════════════════════════════════════════════

def parse_technical_file(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """
    Main parse function for technical question files.

    Returns:
    {
        "sheets": [
            {
                "sheet_name": str,
                "section_name": str,
                "row_count": int,
                "columns_detected": [str],
                "supplier_name": str,
                "questions": [
                    {
                        "question_id": str,
                        "category": str,
                        "question_text": str,
                        "supplier_name": str,
                        "response": str,
                        "comments": str,
                        "compliance_raw": str,
                        "score_hint": float,
                        "status": str,
                        "response_quality": str,
                    },
                    ...
                ]
            },
            ...
        ],
        "total_questions": int,
        "suppliers_detected": [str],
    }
    """

    # Load workbook
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except (InvalidFileException, Exception) as e:
        logger.error(f"Failed to load workbook: {e}")
        raise InvalidFileException(f"File could not be read as Excel: {e}")

    fmt = detect_format(wb)
    result = {"sheets": [], "total_questions": 0, "suppliers_detected": set()}

    for ws in wb.worksheets:
        sections = detect_sections(ws)
        if not sections:
            continue

        supplier_name = extract_supplier_name(ws, ws.title, filename)

        for section in sections:
            # Get header row
            header_row_idx = section["header_row_idx"]
            header_row = [ws.cell(header_row_idx, col_idx + 1).value for col_idx in range(ws.max_column)]

            # Map columns
            col_mapping = map_columns(header_row)

            # Ensure critical columns exist
            if "question_text" not in col_mapping or "response" not in col_mapping:
                logger.warning(f"Sheet {ws.title}, section {section['section_name']}: missing question_text or response column")
                continue

            # Parse data rows
            section_questions = []
            for row_idx in range(section["data_start_idx"], section["data_end_idx"] + 1):
                try:
                    row_values = [ws.cell(row_idx, col_idx + 1).value for col_idx in range(ws.max_column)]

                    # Extract fields
                    qid = None
                    if "question_id" in col_mapping:
                        qid = str(row_values[col_mapping["question_id"]] or "").strip()

                    question_text = str(row_values[col_mapping["question_text"]] or "").strip()
                    if not question_text:
                        continue

                    # Skip stray pricing rows
                    if qid and qid.upper().startswith("SKU"):
                        continue

                    # Generate question ID if not present
                    if not qid:
                        qid = f"{section['section_name'].upper().replace(' ', '_')}-{row_idx}"

                    category = None
                    if "category" in col_mapping:
                        category = str(row_values[col_mapping["category"]] or "").strip()
                    if not category:
                        category = infer_category_from_qid(qid)

                    response = str(row_values[col_mapping["response"]] or "").strip()

                    comments = ""
                    if "doc_ref" in col_mapping:
                        comments = str(row_values[col_mapping["doc_ref"]] or "").strip()

                    compliance_raw = ""
                    compliance_hint = {"score_hint": None, "status": None}
                    if "compliance" in col_mapping:
                        compliance_raw = str(row_values[col_mapping["compliance"]] or "").strip()
                        compliance_hint = compliance_to_score_hint(compliance_raw)

                    response_quality = assess_response_quality(response)

                    question_obj = {
                        "question_id": qid,
                        "category": category,
                        "question_text": question_text,
                        "supplier_name": supplier_name,
                        "response": response,
                        "comments": comments,
                        "compliance_raw": compliance_raw,
                        "score_hint": compliance_hint.get("score_hint"),
                        "status": compliance_hint.get("status"),
                        "response_quality": response_quality.get("quality"),
                    }

                    section_questions.append(question_obj)

                except Exception as e:
                    logger.warning(f"Error parsing row {row_idx} in sheet {ws.title}: {e}")
                    continue

            if section_questions:
                sheet_result = {
                    "sheet_name": ws.title,
                    "section_name": section["section_name"],
                    "row_count": len(section_questions),
                    "columns_detected": list(col_mapping.keys()),
                    "supplier_name": supplier_name,
                    "questions": section_questions,
                }
                result["sheets"].append(sheet_result)
                result["suppliers_detected"].add(supplier_name)

    result["total_questions"] = sum(s["row_count"] for s in result["sheets"])
    result["suppliers_detected"] = sorted(list(result["suppliers_detected"]))

    return result
