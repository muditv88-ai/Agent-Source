"""
pricing_parser.py v3.1

Phase 1 (AI): Parse the pricing/commercial sheet, understand its structure,
              identify buyer-defined fields vs supplier-filled fields.
Phase 2 (Python): Return a structured PricingSheet object ready for
                  pure-Python comparison and scenario calculation.

Supports:
  - Excel (.xlsx/.xls), CSV, PDF/text with tables
  - Simple unit-price model  (Unit Price × Qty = Total)
  - Cost-breakdown model     (API Cost + RM + Pkg + Mfg + Overhead + Margin = Unit Total)
  - Rate-card, flat-rate, category-based structures
  - LLM fallback when structural parse yields no prices (up to 12k chars)

v3 changes:
  - Cost-breakdown column detection (API Cost, RM Cost, Pkg Cost, Mfg Cost, Overhead, Margin)
  - Expanded _find() synonym list to match real-world supplier column names
  - Section-header skip logic (avoids treating SECTION A/B/C rows as data)
  - rfp_full_text now injected into LLM prompts for column context
  - LLM fallback text limit raised from 8k → 12k chars
  - Graceful handling of pharmaceutical SKU# / Drug Name / Strength combos

v3.1: added parse_pricing_response as a backward-compat alias for
      extract_pricing_from_document so that pricing_agent.py can import
      either name without error.
"""
import os
import re
import json
import time
from typing import Any
from openai import OpenAI

_client = OpenAI(
    base_url="https://integrate.api.nvidia.com/v1",
    api_key=os.environ.get("NVIDIA_API_KEY"),
)
MODEL = "meta/llama-3.3-70b-instruct"

# ── Structure labels ──────────────────────────────────────────────────────────
STRUCTURE_LINE_ITEM   = "line_item"
STRUCTURE_FLAT_RATE   = "flat_rate"
STRUCTURE_CATEGORY    = "category_based"
STRUCTURE_RATE_CARD   = "rate_card"
STRUCTURE_MIXED       = "mixed"
STRUCTURE_COST_BREAKDOWN = "cost_breakdown"   # NEW — pharma / manufacturing model

# ── Section-header patterns to skip ──────────────────────────────────────────
_SECTION_SKIP_RE = re.compile(
    r"^(SECTION [A-Z]|PART [A-Z0-9]|APPENDIX|EXHIBIT|ATTACHMENT|SCHEDULE)\b",
    re.IGNORECASE,
)

# ── Cost-breakdown column groups (v3) ─────────────────────────────────────────
# Each tuple: (output_key, list_of_header_substrings_to_match)
_COST_BREAKDOWN_COLS = [
    ("api_cost",   ["api cost", "api", "active ingredient", "raw material cost", "rm cost", "material cost"]),
    ("rm_cost",    ["rm cost", "raw mat", "excipient", "other material"]),
    ("pkg_cost",   ["pkg cost", "pack cost", "packaging", "container", "label"]),
    ("mfg_cost",   ["mfg cost", "manufacturing cost", "conversion", "production cost", "process cost"]),
    ("overhead",   ["overhead", "indirect", "burden", "opex"]),
    ("margin",     ["margin", "profit", "markup", "mark-up", "gp"]),
]

# ── RFP structure understanding prompt ───────────────────────────────────────
_RFP_STRUCTURE_PROMPT = """
You are a procurement data analyst. Analyse the pricing/commercial section below.

Identify:
1. BUYER-DEFINED fields: values pre-filled by the buyer in the RFP template
   (e.g. SKU codes, item descriptions, quantities, specifications, target prices)
2. SUPPLIER-FILLED fields: columns/cells the supplier was asked to complete
   (e.g. unit price, total cost, discount %, delivery lead time, payment terms,
    API cost, RM cost, packaging cost, manufacturing cost, overhead, margin)
3. The pricing structure type:
   line_item | flat_rate | category_based | rate_card | mixed | cost_breakdown
   Use cost_breakdown when the supplier provides a cost build-up
   (e.g. API Cost + RM Cost + Packaging + Manufacturing + Overhead + Margin = Unit Total)
4. The currency (if detectable)
5. Whether there is a Grand Total / Total Cost / Unit Total cell

Return ONLY this JSON:
{
  "structure_type": "cost_breakdown",
  "currency": "USD",
  "buyer_fields": ["SKU#", "Drug Name", "Strength", "Dosage Form", "Annual Vol"],
  "supplier_fields": ["API Cost", "RM Cost", "Pkg Cost", "Mfg Cost", "Overhead", "Margin", "Unit Total"],
  "has_grand_total": false,
  "notes": "any important structural observations"
}
"""

# ── LLM line-item extraction prompt ──────────────────────────────────────────
_LLM_PRICING_PROMPT = """
You are a procurement data extraction assistant.
Extract ALL pricing line items from the supplier document text below.

IMPORTANT: This document may use a COST BREAKDOWN model where the unit price
is built up from components: API Cost + RM Cost + Pkg Cost + Mfg Cost + Overhead + Margin = Unit Total.
In that case, set unit_price = Unit Total (the final all-in price per unit).
Also capture individual cost components in the "notes" field if present.

Return ONLY a JSON array. Each item must have:
- "sku": string (SKU/part number if present, else "")
- "description": string (item/service name — required; for pharma: "Drug Name Strength Dosage Form")
- "quantity": number (annual volume if stated, else 1)
- "unit_price": number (final unit total / all-in price per unit, 0 if blank)
- "total": number (quantity × unit_price, or stated total)
- "category": string (infer from drug class or section: Cardiovascular / Antibiotics / Analgesics / Respiratory / etc.)
- "unit": string (tablet / capsule / vial / inhaler / syringe — if stated, else "unit")
- "is_buyer_defined": true if SKU/description was pre-filled by buyer template
- "notes": string (cost breakdown details, discounts, pack type, conditions)

Rules:
- Include ALL items with a price. Skip items where unit_price AND total are both 0.
- Strip currency symbols ($, £, €, ₹) and commas from numbers.
- If only total given and qty>1: unit_price = total / qty.
- For cost-breakdown rows: unit_price = the "Unit Total" or rightmost total column.
- Return [] if truly no pricing found.

Document text (up to 12000 chars):
"""


# ── LLM helpers ───────────────────────────────────────────────────────────────

def _call_llm(prompt: str, max_tokens: int = 1024, max_retries: int = 3) -> str:
    delay = 15.0
    for attempt in range(max_retries + 1):
        try:
            resp = _client.chat.completions.create(
                model=MODEL,
                messages=[
                    {"role": "system", "content": "detailed thinking off"},
                    {"role": "user",   "content": prompt},
                ],
                temperature=0.1,
                max_tokens=max_tokens,
            )
            return resp.choices[0].message.content or ""
        except Exception as e:
            msg = str(e).lower()
            if ("429" in msg or "rate limit" in msg) and attempt < max_retries:
                time.sleep(delay)
                delay *= 2
                continue
            raise
    return ""


def _parse_json(raw: str) -> Any:
    raw = raw.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s*```$", "", raw).strip()
    return json.loads(raw)


def _llm_understand_structure(text: str, rfp_text: str = "") -> dict:
    """Ask LLM to identify buyer vs supplier fields."""
    context = ""
    if rfp_text:
        context = "\n\nRFP template column definitions (buyer-defined):\n" + rfp_text[:1500]
    prompt = _RFP_STRUCTURE_PROMPT + context + "\n\nSupplier document text (up to 3000 chars):\n" + text[:3000]
    try:
        raw = _call_llm(prompt, max_tokens=512)
        return _parse_json(raw)
    except Exception:
        return {
            "structure_type": "line_item",
            "currency": "",
            "buyer_fields": [],
            "supplier_fields": [],
            "has_grand_total": False,
            "notes": "Structure auto-detection failed",
        }


def _llm_extract_line_items(text: str, supplier_name: str, rfp_text: str = "") -> list:
    """LLM fallback: extract structured line items from unstructured text."""
    context = ""
    if rfp_text:
        context = "\nRFP column context:\n" + rfp_text[:1000] + "\n"
    prompt = _LLM_PRICING_PROMPT + context + text[:12000]
    try:
        raw   = _call_llm(prompt, max_tokens=4096)
        items = _parse_json(raw)
        if not isinstance(items, list):
            items = items.get("line_items") or items.get("items") or []
        return _sanitise_items(items)
    except Exception:
        return []


def _sanitise_items(items: list) -> list:
    clean = []
    for item in items:
        try:
            qty        = float(item.get("quantity", 1) or 1)
            unit_price = float(item.get("unit_price", 0) or 0)
            total      = float(item.get("total", 0) or 0)
            if total == 0 and unit_price > 0:
                total = round(qty * unit_price, 2)
            if unit_price == 0 and total > 0:
                unit_price = round(total / qty, 2)
            if total == 0 and unit_price == 0:
                continue
            clean.append({
                "sku":              str(item.get("sku", "")).strip(),
                "description":      str(item.get("description", "")).strip(),
                "quantity":         qty,
                "unit_price":       unit_price,
                "total":            total,
                "category":         str(item.get("category", "General")).strip(),
                "unit":             str(item.get("unit", "each")).strip(),
                "is_buyer_defined": bool(item.get("is_buyer_defined", False)),
                "notes":            str(item.get("notes", "")).strip(),
            })
        except (TypeError, ValueError):
            continue
    return clean


# ── Number cleaner ────────────────────────────────────────────────────────────

def _clean_number(val: Any) -> float | None:
    if val is None:
        return None
    s = re.sub(r"[$£€₹%,]", "", str(val).strip())
    try:
        return float(s)
    except ValueError:
        return None


# ── Excel / CSV / text parsers ────────────────────────────────────────────────

def _parse_excel_bytes(content: bytes) -> dict:
    import openpyxl, io
    wb     = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            continue
        # Find header row: first row with ≥2 non-empty cells and at least one
        # cell that looks like a column name (non-numeric)
        hdr_idx = 0
        for i, row in enumerate(rows_raw):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            has_label = any(
                not re.match(r'^[\d\$\€\£\₹\.\,\%\-]+$', str(c).strip())
                for c in non_empty
            )
            if len(non_empty) >= 2 and has_label:
                hdr_idx = i
                break
        headers   = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows_raw[hdr_idx])]
        data_rows = [
            dict(zip(headers, row))
            for row in rows_raw[hdr_idx + 1:]
            if not all(c is None for c in row)
        ]
        if data_rows:
            sheets.append({"sheet_name": sheet_name, "headers": headers, "rows": data_rows})
    return {"source": "excel", "sheets": sheets}


def _parse_csv_text(text: str) -> dict:
    import csv, io
    reader  = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])
    return {"source": "csv", "sheets": [{"sheet_name": "Sheet1", "headers": headers, "rows": list(reader)}]}


def _extract_tables_from_text(text: str) -> dict:
    lines = text.split("\n")
    blocks, current = [], []
    for line in lines:
        s = line.strip()
        if "|" in s or re.search(r"\d+[\.,]\d+.*\d+[\.,]\d+", s):
            current.append(s)
        else:
            if len(current) >= 3:
                blocks.append(current)
            current = []
    if len(current) >= 3:
        blocks.append(current)

    sheets = []
    for idx, block in enumerate(blocks):
        if any("|" in l for l in block):
            parsed = [[c.strip() for c in l.split("|") if c.strip()] for l in block]
            parsed = [r for r in parsed if r]
        else:
            parsed = [re.split(r"\s{2,}", l.strip()) for l in block]
            parsed = [r for r in parsed if len(r) >= 2]
        if len(parsed) >= 2:
            hdrs = parsed[0]
            rows = [dict(zip(hdrs, r)) for r in parsed[1:]]
            sheets.append({"sheet_name": f"Table_{idx+1}", "headers": hdrs, "rows": rows})
    return {"source": "text", "sheets": sheets}


# ── Structure detector ────────────────────────────────────────────────────────

def _detect_structure(headers: list) -> str:
    h = " ".join(headers).lower()
    # Cost breakdown: has multiple cost component columns
    breakdown_hits = sum(
        1 for _, synonyms in _COST_BREAKDOWN_COLS
        if any(s in h for s in synonyms)
    )
    if breakdown_hits >= 3:
        return STRUCTURE_COST_BREAKDOWN
    has_sku        = any(k in h for k in ["sku", "item", "product", "part", "line", "description", "drug", "material"])
    has_qty        = any(k in h for k in ["qty", "quantity", "units", "volume", "annual vol", "packs"])
    has_unit_price = any(k in h for k in [
        "unit price", "unit cost", "unit total", "rate", "price per", "cost per",
        "quoted price", "offered price", "vendor price", "supplier price",
        "net price", "sell price", "list price", "total price", "amount",
    ])
    has_category   = any(k in h for k in ["category", "section", "group", "type", "drug class", "therapeutic"])
    has_role       = any(k in h for k in ["role", "resource", "level", "grade", "tier", "day rate", "hourly"])
    has_flat       = any(k in h for k in ["one-time", "setup", "annual", "monthly", "recurring", "subscription", "licence", "license"])
    if has_role and has_unit_price:     return STRUCTURE_RATE_CARD
    if has_flat and not has_sku:        return STRUCTURE_FLAT_RATE
    if has_category and not has_sku:    return STRUCTURE_CATEGORY
    if has_sku and has_unit_price:      return STRUCTURE_LINE_ITEM
    return STRUCTURE_MIXED


# ── Sheet normaliser ──────────────────────────────────────────────────────────

def _normalise_sheet(sheet: dict, supplier_name: str, buyer_fields: list, supplier_fields: list) -> dict:
    headers  = sheet.get("headers", [])
    rows     = sheet.get("rows", [])
    h_lower  = [h.lower() for h in headers]
    structure = _detect_structure(headers)

    buyer_cols    = [h for h in headers if any(bf.lower() in h.lower() for bf in buyer_fields)]
    supplier_cols = [h for h in headers if any(sf.lower() in h.lower() for sf in supplier_fields)]

    def _find(*candidates):
        """Return header name for first candidate substring found."""
        for c in candidates:
            for i, h in enumerate(h_lower):
                if c in h:
                    return headers[i]
        return None

    # ── Column mapping — expanded synonyms for real-world supplier docs ──────
    sku_col = _find(
        "sku", "sku#", "part", "code", "item no", "item#", "item code",
        "product code", "article", "ref", "material no",
    )
    desc_col = _find(
        "description", "item", "product", "service", "name", "drug name",
        "drug", "medicine", "material", "role", "resource", "category", "section",
    )
    strength_col = _find("strength", "dose", "concentration", "potency")
    form_col     = _find("dosage form", "form", "formulation", "type")
    qty_col = _find(
        "annual vol", "qty", "quantity", "units", "volume", "hours", "days",
        "packs", "annual quantity", "annual units",
    )
    # Price column: try specific then generic, prefer "unit total" for cost-breakdown
    price_col = _find(
        "unit total", "total unit", "all-in", "all in price",
        "unit price", "unit cost", "rate", "quoted price", "offered price",
        "vendor price", "supplier price", "net price", "sell price", "list price",
        "price", "cost", "fee", "amount",
    )
    total_col  = _find("total cost", "total", "extended", "subtotal", "line total", "grand total")
    cat_col    = _find("drug class", "therapeutic", "category", "group", "section", "type")
    unit_col   = _find("dosage form", "unit", "uom", "measure", "form")
    notes_col  = _find("comments", "notes", "remarks", "conditions", "caveats")
    pack_col   = _find("pack type", "pack", "packaging", "container")

    # ── Cost-breakdown component columns ─────────────────────────────────────
    breakdown_col_map = {}
    if structure == STRUCTURE_COST_BREAKDOWN:
        for key, synonyms in _COST_BREAKDOWN_COLS:
            col = _find(*synonyms)
            if col:
                breakdown_col_map[key] = col

    line_items = []
    for row in rows:
        # Skip section-header rows
        first_val = str(list(row.values())[0] or "").strip()
        if _SECTION_SKIP_RE.match(first_val):
            continue

        sku  = str(row.get(sku_col, "")).strip()  if sku_col  else ""
        desc = str(row.get(desc_col, "")).strip() if desc_col else ""

        # For pharma docs: build description from Drug Name + Strength + Form
        if strength_col or form_col:
            parts = [desc]
            if strength_col:
                s = str(row.get(strength_col, "")).strip()
                if s and s not in desc:
                    parts.append(s)
            if form_col:
                f = str(row.get(form_col, "")).strip()
                if f and f not in desc:
                    parts.append(f)
            desc = " ".join(p for p in parts if p).strip()

        qty        = _clean_number(row.get(qty_col))    if qty_col   else 1.0
        unit_price = _clean_number(row.get(price_col)) if price_col else None
        total      = _clean_number(row.get(total_col)) if total_col else None
        category   = str(row.get(cat_col, "")).strip()  if cat_col   else sheet.get("sheet_name", "")
        unit       = str(row.get(unit_col, "each")).strip() if unit_col else "each"
        notes_raw  = str(row.get(notes_col, "")).strip() if notes_col else ""
        pack       = str(row.get(pack_col, "")).strip()  if pack_col  else ""

        # Build notes with cost breakdown detail
        notes_parts = []
        if pack:
            notes_parts.append(f"Pack: {pack}")
        if breakdown_col_map:
            bd_parts = []
            for key, col in breakdown_col_map.items():
                v = _clean_number(row.get(col))
                if v is not None and v > 0:
                    bd_parts.append(f"{key}={v:.4f}")
            if bd_parts:
                notes_parts.append("Cost breakdown: " + ", ".join(bd_parts))
        if notes_raw:
            notes_parts.append(notes_raw)
        notes = " | ".join(notes_parts)

        is_buyer_defined = bool(desc_col and desc_col in buyer_cols)

        if not desc and unit_price is None and total is None:
            continue
        if total is None and unit_price is not None and qty:
            total = round((unit_price) * (qty or 1), 2)
        if unit_price is None and total is not None and qty:
            unit_price = round(total / (qty or 1), 2)

        # Skip rows with no numeric pricing at all
        if (unit_price is None or unit_price == 0) and (total is None or total == 0):
            continue

        line_items.append({
            "sku":             sku,
            "description":     desc,
            "quantity":        qty or 1.0,
            "unit_price":      unit_price or 0.0,
            "total":           total or 0.0,
            "category":        category,
            "unit":            unit,
            "is_buyer_defined": is_buyer_defined,
            "notes":           notes,
        })

    return {
        "sheet_name":     sheet.get("sheet_name", "Pricing"),
        "structure_type": structure,
        "supplier_name":  supplier_name,
        "headers":        headers,
        "buyer_cols":     buyer_cols,
        "supplier_cols":  supplier_cols,
        "line_items":     line_items,
    }


# ── Main entry point ──────────────────────────────────────────────────────────

def extract_pricing_from_document(
    file_path: str,
    supplier_name: str,
    full_text: str = "",
    rfp_full_text: str = "",
) -> dict:
    """
    Two-phase extraction:
    Phase 1 — LLM analyses structure (buyer vs supplier fields).
    Phase 2 — Python parses the actual data using that structure map.

    Returns:
        supplier_name, structure_type, structure_info (LLM analysis),
        sheets, all_line_items, total_cost, source_format, parse_warnings
    """
    ext = os.path.splitext(file_path)[1].lower()
    raw = {"source": "unknown", "sheets": []}

    # --- Structural parse ---
    try:
        if ext in (".xlsx", ".xls"):
            with open(file_path, "rb") as f:
                raw = _parse_excel_bytes(f.read())
        elif ext == ".csv":
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                raw = _parse_csv_text(f.read())
        else:
            if full_text:
                raw = _extract_tables_from_text(full_text)
    except Exception as e:
        raw = {"source": "error", "sheets": [], "error": str(e)}

    # --- Phase 1: LLM structure understanding ---
    sample_text = full_text[:3000] if full_text else ""
    if not sample_text and raw.get("sheets"):
        first = raw["sheets"][0]
        sample_text = " | ".join(first.get("headers", [])) + "\n"
        for row in list(first.get("rows", []))[:10]:
            sample_text += " | ".join(str(v) for v in row.values() if v is not None) + "\n"

    structure_info = _llm_understand_structure(sample_text, rfp_full_text) if sample_text else {
        "structure_type": "line_item", "currency": "",
        "buyer_fields": [], "supplier_fields": [], "has_grand_total": False, "notes": ""
    }
    buyer_fields    = structure_info.get("buyer_fields", [])
    supplier_fields = structure_info.get("supplier_fields", [])

    # --- Phase 2: Python normalisation ---
    normalised_sheets = [
        _normalise_sheet(sheet, supplier_name, buyer_fields, supplier_fields)
        for sheet in raw.get("sheets", [])
        if sheet.get("rows")
    ]
    if not normalised_sheets and full_text:
        text_raw = _extract_tables_from_text(full_text)
        normalised_sheets = [
            _normalise_sheet(sheet, supplier_name, buyer_fields, supplier_fields)
            for sheet in text_raw.get("sheets", [])
            if sheet.get("rows")
        ]

    all_items    = [item for sheet in normalised_sheets for item in sheet["line_items"]]
    total_priced = sum(i["total"] for i in all_items)
    used_llm     = False

    # --- LLM fallback if structural parse yielded nothing ---
    if total_priced == 0 and full_text:
        llm_items = _llm_extract_line_items(full_text, supplier_name, rfp_full_text)
        if llm_items:
            all_items    = llm_items
            total_priced = sum(i["total"] for i in llm_items)
            used_llm     = True
            normalised_sheets = [{
                "sheet_name":     "LLM Extracted",
                "structure_type": "mixed",
                "supplier_name":  supplier_name,
                "headers":        ["sku", "description", "quantity", "unit_price", "total", "category", "unit", "notes"],
                "buyer_cols":     buyer_fields,
                "supplier_cols":  supplier_fields,
                "line_items":     llm_items,
            }]

    structure_types = list({s["structure_type"] for s in normalised_sheets})
    warnings = []
    if total_priced == 0:
        warnings.append("No pricing detected — use the Correct button to enter prices manually")
    if used_llm:
        warnings.append("Prices extracted by AI — please verify values in the Price Matrix tab")

    return {
        "supplier_name":   supplier_name,
        "structure_type":  structure_types[0] if len(structure_types) == 1 else "mixed",
        "structure_info":  structure_info,
        "sheets":          normalised_sheets,
        "all_line_items":  all_items,
        "total_cost":      round(total_priced, 2),
        "source_format":   raw.get("source", "unknown"),
        "parse_warnings":  warnings,
        "currency":        structure_info.get("currency", ""),
    }


# ── backward-compat alias ─────────────────────────────────────────────────────
# pricing_agent.py imports `parse_pricing_response`; this alias satisfies that
# import without changing any existing callers of extract_pricing_from_document.
def parse_pricing_response(file_text: str, supplier_name: str = "Supplier", **kwargs) -> dict:
    """
    Alias shim for pricing_agent.py compatibility.
    Accepts a text string instead of a file path by writing to a temp file,
    then delegates to extract_pricing_from_document.
    """
    import tempfile
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".txt", delete=False, encoding="utf-8"
    ) as tmp:
        tmp.write(file_text)
        tmp_path = tmp.name
    try:
        return extract_pricing_from_document(
            file_path=tmp_path,
            supplier_name=supplier_name,
            full_text=file_text,
            rfp_full_text=kwargs.get("rfp_full_text", ""),
        )
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

# ═══ parse_pricing_sheet_with_diagnostics v1 ══════════════════════════════════
import re as _re2, io as _io2, openpyxl as _openpyxl
_EXCL=["request for proposal","instructions:","shaded columns","annual contract value =","title ","section ","part a","part b","appendix","exhibit","attachment","schedule","note:","notes:","please fill","supplier to complete","buyer to complete","highlighted cells","yellow cells","orange cells","do not modify","do not change","for internal use","page ","confidential","rfp template","commercial price sheet","technical price sheet","pricing sheet","cost breakdown"]
_IKR=_re2.compile(r"^(SKU|ITEM|PART|PROD|REF|MAT|API|NDC)[-_ ]?[0-9A-Z]",_re2.IGNORECASE)
_NKR=_re2.compile(r"^[0-9]{1,5}([.\-][0-9])?$")
_HW={"sku","item","part","description","qty","quantity","price","cost","total","unit","volume","annual","drug","strength","form","supplier","buyer"}
_PS=["commercial","pricing","price","cost","quotation","quote","unit price","rate card","costing"]
_HS=["sku","item","description","drug","qty","quantity","annual","unit price","unit cost","total","cost","price","volume","strength","form","margin","overhead","packaging"]
_CN={"sku":"item_key","sku#":"item_key","item no":"item_key","item #":"item_key","part":"item_key","material no":"item_key","description":"description","drug name":"description","item":"description","product":"description","service":"description","annual vol":"annual_volume","annual volume":"annual_volume","qty":"quantity","quantity":"quantity","units":"quantity","api cost":"api_cost","active ingredient":"api_cost","raw material cost":"rm_cost","rm cost":"rm_cost","packaging cost":"pkg_cost","pkg cost":"pkg_cost","manufacturing cost":"mfg_cost","mfg cost":"mfg_cost","overhead cost":"overhead","overhead":"overhead","profit margin":"margin","margin":"margin","total unit cost":"total_unit_cost","unit total":"total_unit_cost","unit price":"total_unit_cost","all-in":"total_unit_cost","quoted price":"total_unit_cost","net transfer price":"total_unit_cost","all in cost":"total_unit_cost","annual contract value":"annual_contract_value","annual value":"annual_contract_value","currency":"currency","uom":"uom","unit":"uom","lead time":"lead_time","moq":"moq","shelf life":"shelf_life","storage":"storage_condition"}

def _vr(rv,ik,mn):
    j=" ".join(str(v) for v in rv if v is not None).strip().lower()
    for p in _EXCL:
        if p in j: return False,f"excluded_phrase:{p!r}"
    if ik:
        k=str(ik).strip()
        if not (_IKR.match(k) or _NKR.match(k)):
            if k.lower() in _HW: return False,f"repeated_header:{k!r}"
    ne=sum(1 for v in rv if v is not None and str(v).strip())
    m=max(3,mn//3) if mn else 3
    if ne<m: return False,f"sparse:{ne}<{m}"
    return True,""

def parse_pricing_sheet_with_diagnostics(file_bytes,file_ext,supplier_name):
    wb=_openpyxl.load_workbook(_io2.BytesIO(file_bytes),data_only=True)
    bw,bs=None,-1
    for ws in wb.worksheets:
        sc=sum(3 for s in _PS if s in ws.title.lower())
        c=t=0
        for row in ws.iter_rows(max_row=20,values_only=True):
            for x in row:
                if x is not None:
                    t+=1
                    try: float(str(x).replace(",","").replace("$","")); c+=1
                    except: pass
        if t: sc+=int(c/t*10)
        if sc>bs: bs,bw=sc,ws
    ws=bw or wb.active
    hi,hds=None,[]
    for i,row in enumerate(ws.iter_rows(max_row=15,values_only=True)):
        txt=" ".join(str(c).lower() for c in row if c is not None)
        if sum(1 for s in _HS if s in txt)>=3: hi=i; hds=[str(c).strip() if c else "" for c in row]; break
    if hi is None:
        hi=0; hds=[str(c).strip() if c else "" for c in list(ws.iter_rows(max_row=1,values_only=True))[0]]
    cm={}
    hl=[h.lower().strip() for h in hds]
    for i,h in enumerate(hl):
        if not h: continue
        if h in _CN: cm[hds[i]]=_CN[h]; continue
        for k,v in _CN.items():
            if k in h: cm[hds[i]]=v; break
    ii=next((i for i,h in enumerate(hds) if cm.get(h)=="item_key"),0)
    acc,exc=[],[]
    ar=list(ws.iter_rows(min_row=hi+2,values_only=True))
    rn=sum(1 for r in ar if any(c is not None and str(c).strip() for c in r))
    for row in ar:
        if not any(c is not None and str(c).strip() for c in row): continue
        rv=list(row); ik=rv[ii] if ii<len(rv) else None
        ok,reason=_vr(rv,ik,len(cm))
        rd={hds[i]:rv[i] for i in range(min(len(hds),len(rv)))}
        if ok: acc.append(rd)
        else:
            p=" | ".join(str(v) for v in rv[:6] if v is not None)[:120]
            exc.append({"reason":reason,"preview":p})
    mc=set(cm.values())
    hk="item_key" in mc or "description" in mc
    ht="total_unit_cost" in mc or "annual_contract_value" in mc
    hq="annual_volume" in mc or "quantity" in mc
    cf="high" if(hk and ht and hq) else "medium" if(hk and(ht or hq)) else "low"
    w=[]
    if not ht: w.append("No total cost column detected")
    if not acc: w.append("No line items accepted"); cf="low"
    if len(exc)>len(acc): w.append(f"More excluded ({len(exc)}) than accepted ({len(acc)})")
    return {"detected_sheet_name":ws.title,"raw_non_empty_rows":rn,"accepted_line_items":len(acc),"accepted_line_items_data":acc,"excluded_rows":exc,"column_mapping":cm,"sample_rows":acc[:5],"parse_confidence":cf,"warnings":w}


# ══════════════════════════════════════════════════════════════════════════════
# parse_pricing_sheet_with_diagnostics  — v1
# ══════════════════════════════════════════════════════════════════════════════
import re as _re_ps, io as _io_ps

_PRICING_SHEET_KEYWORDS = [
    "price", "pricing", "cost", "quotation", "quote", "rate card",
    "costing", "commercial", "unit price", "unit cost",
]
_HEADER_SYNONYMS = {
    "sku": "item_key", "sku#": "item_key", "sku no": "item_key",
    "item": "item_key", "item no": "item_key", "item #": "item_key",
    "part": "item_key", "part no": "item_key", "material no": "item_key",
    "description": "description", "drug name": "description",
    "product": "description", "service": "description",
    "annual vol": "annual_volume", "annual volume": "annual_volume",
    "qty": "quantity", "quantity": "quantity", "units": "quantity",
    "api cost": "api_cost", "active ingredient cost": "api_cost",
    "raw material": "raw_material_cost", "raw material cost": "raw_material_cost",
    "packaging": "packaging_cost", "packaging cost": "packaging_cost",
    "manufacturing": "manufacturing_cost", "manufacturing cost": "manufacturing_cost",
    "overhead": "overhead_cost", "overhead cost": "overhead_cost",
    "margin": "margin", "profit margin": "margin",
    "total unit cost": "total_unit_cost", "unit price": "total_unit_cost",
    "total cost": "total_unit_cost", "net price": "total_unit_cost",
    "annual contract value": "annual_contract_value",
    "annual value": "annual_contract_value", "extended value": "annual_contract_value",
    "currency": "currency", "moq": "moq", "lead time": "lead_time",
    "shelf life": "shelf_life", "storage": "storage_condition",
    "uom": "uom", "unit of measure": "uom",
}
_SKIP_PATTERNS = [
    _re_ps.compile(p, _re_ps.IGNORECASE) for p in [
        r"instructions?:", r"please (fill|enter|provide)",
        r"shaded columns?", r"annual contract value\s*=",
        r"^(notes?|note):?$", r"^(section|part)\s+[a-z0-9]",
        r"^(appendix|exhibit|attachment|schedule)\b",
        r"supplier to complete", r"buyer to complete",
        r"do not (modify|change)", r"for internal use",
        r"confidential", r"rfp template",
        r"^page\s+\d+", r"highlighted cells", r"yellow cells",
    ]
]


def _is_skip_row(row_text: str) -> bool:
    t = row_text.strip()
    if not t:
        return True
    return any(p.search(t) for p in _SKIP_PATTERNS)


def _score_sheet(sheet_name: str, rows: list) -> int:
    score = 0
    name_lower = sheet_name.lower()
    for kw in _PRICING_SHEET_KEYWORDS:
        if kw in name_lower:
            score += 10
    sample_text = " ".join(
        str(c) for row in rows[:20] for c in row if c is not None
    ).lower()
    for kw in _PRICING_SHEET_KEYWORDS:
        if kw in sample_text:
            score += 3
    numeric_cells = sum(
        1 for row in rows[:30] for c in row
        if isinstance(c, (int, float)) and c not in (0, None)
    )
    score += min(numeric_cells, 20)
    return score


def _detect_header_row(rows: list) -> tuple[int, dict]:
    """Return (header_row_index, column_mapping)."""
    for i, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        matched = sum(1 for c in cells if c in _HEADER_SYNONYMS)
        if matched >= 2:
            mapping = {}
            for j, cell in enumerate(cells):
                if cell in _HEADER_SYNONYMS:
                    mapping[j] = _HEADER_SYNONYMS[cell]
            return i, mapping
    return -1, {}


def _confidence(mapping: dict, rows_count: int) -> str:
    canonical_hits = len(set(mapping.values()))
    has_total = any(v in ("total_unit_cost", "annual_contract_value") for v in mapping.values())
    has_key   = any(v == "item_key" for v in mapping.values())
    if canonical_hits >= 4 and has_total and has_key and rows_count >= 3:
        return "high"
    if canonical_hits >= 2 and rows_count >= 1:
        return "medium"
    return "low"


def parse_pricing_sheet_with_diagnostics(
    file_bytes: bytes,
    file_ext: str,
    supplier_name: str = "Unknown",
) -> dict:
    import openpyxl as _xl

    warnings: list[str] = []

    # ── load workbook ────────────────────────────────────────────────────────
    try:
        wb = _xl.load_workbook(_io_ps.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open workbook: {e}")

    # ── pick best sheet ──────────────────────────────────────────────────────
    best_sheet, best_score = wb.worksheets[0], -1
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        score = _score_sheet(ws.title, rows)
        if score > best_score:
            best_score, best_sheet = score, ws

    all_rows = list(best_sheet.iter_rows(values_only=True))
    non_empty = [r for r in all_rows if any(c is not None for c in r)]

    # ── detect header row ────────────────────────────────────────────────────
    header_idx, col_map = _detect_header_row(non_empty)
    if header_idx == -1:
        warnings.append("Could not detect a header row — using column positions as keys")
        header_idx = 0
        col_map = {i: f"col_{i}" for i in range(len(non_empty[0])) if non_empty[0][i] is not None}

    human_col_map = {
        str(non_empty[header_idx][k]) if non_empty[header_idx][k] is not None else f"col_{k}": v
        for k, v in col_map.items()
    }

    # ── extract data rows ────────────────────────────────────────────────────
    data_rows = non_empty[header_idx + 1:]
    accepted, excluded = [], []

    for row in data_rows:
        row_text = " ".join(str(c) for c in row if c is not None)
        if _is_skip_row(row_text):
            excluded.append({"raw": row_text[:120], "reason": "instruction/header/blank row"})
            continue
        has_numeric = any(isinstance(row[i], (int, float)) for i in col_map if i < len(row))
        has_text    = any(isinstance(row[i], str) and row[i].strip() for i in col_map if i < len(row))
        if not (has_numeric or has_text):
            excluded.append({"raw": row_text[:120], "reason": "no usable content"})
            continue
        record = {}
        for col_idx, canonical in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[canonical] = val
        accepted.append(record)

    if not accepted:
        warnings.append("No data rows extracted — sheet may be template-only or heavily instruction-based")

    sample = accepted[:5]
    confidence = _confidence(col_map, len(accepted))

    return {
        "supplier_name":            supplier_name,
        "detected_sheet_name":      best_sheet.title,
        "raw_non_empty_rows":       len(non_empty),
        "accepted_line_items":      len(accepted),
        "accepted_line_items_data": accepted,
        "excluded_rows":            excluded[:20],
        "column_mapping":           human_col_map,
        "sample_rows":              sample,
        "parse_confidence":         confidence,
        "warnings":                 warnings,
    }
