"""
pricing_agent.py  v2.0

FM-7.1  Parse + normalize pricing sheets (UoM normalization)
FM-7.2  Full pricing analysis via PricingAgent
FM-7.3  TCO calculator  (unit + freight + duty + tooling)
FM-7.4  Currency normalization (live FX)
FM-7.5  Price validity tracking (expired / expiring soon)
FM-7.6  Side-by-side comparison output

NEW in v2.0:
  - Integrates pricing_sheet_classifier for sheet detection
  - Integrates pricing_schema_mapper for canonical schema output
  - Integrates pricing_validation for confidence-tiered validation
  - ingest_workbook() — primary entry point for Excel/XLSX supplier responses
  - Structured supplier response ingestion with missing-field auto-detection
  - Backward compatible: parse_and_normalize and run() still work as before
"""
import io
import logging
from datetime import datetime, date
from typing import Any, Dict, List, Optional

from app.agents.base_agent import BaseAgent, Tool
from app.services.pricing_analyzer import build_cost_model, analyze_pricing
from app.services.pricing_parser import parse_pricing_response
from app.services.pricing_sheet_classifier import (
    classify_workbook_sheets,
    get_best_pricing_sheet,
    classify_sheet,
)
from app.services.pricing_schema_mapper import map_rows_to_schema
from app.services.pricing_validation import validate_pricing_schema

logger = logging.getLogger(__name__)

# ── UoM conversion table ──────────────────────────────────────────────────────
UOM_FACTORS: Dict[str, float] = {
    "each": 1.0, "unit": 1.0, "pcs": 1.0,
    "kg": 1.0, "g": 0.001, "ton": 1000.0, "mt": 1000.0,
    "l": 1.0, "ml": 0.001,
    "m": 1.0, "cm": 0.01, "mm": 0.001,
    "box_of_10": 10.0, "box_of_100": 100.0,
    "dozen": 12.0, "pair": 2.0, "set": 1.0,
    "hour": 1.0, "day": 8.0,
    # pharma-specific
    "tablet": 1.0, "capsule": 1.0, "vial": 1.0,
    "syringe": 1.0, "inhaler": 1.0, "sachet": 1.0,
    "ampoule": 1.0, "pen": 1.0,
}


class PricingAgent(BaseAgent):
    """
    Orchestrates the full pricing intelligence pipeline.
    Wraps existing pricing_analyzer + pricing_parser + new classifier/mapper.
    """

    def __init__(self, base_currency: str = "USD"):
        self.base_currency = base_currency

        tools = [
            Tool(
                name="ingest_workbook",
                description="Ingest an Excel workbook bytes — detect pricing sheets, map columns, validate",
                fn=self._ingest_workbook,
                schema={"type": "object"},
            ),
            Tool(
                name="parse_and_normalize",
                description="Parse pricing sheet text and normalize units of measure",
                fn=self._parse_normalize,
                schema={"type": "object"},
            ),
            Tool(
                name="calculate_tco",
                description="Calculate Total Cost of Ownership including freight, duty and tooling",
                fn=self._calculate_tco,
                schema={"type": "object"},
            ),
            Tool(
                name="normalize_currency",
                description="Convert all prices to a common base currency using live FX rates",
                fn=self._normalize_currency,
                schema={"type": "object"},
            ),
            Tool(
                name="check_price_validity",
                description="Flag expired or near-expiry supplier quotes",
                fn=self._check_validity,
                schema={"type": "object"},
            ),
        ]
        super().__init__(tools)

    # ── Primary entry point: workbook bytes ──────────────────────────────────

    def _ingest_workbook(
        self,
        file_bytes: bytes,
        supplier_name: str = "",
        source_type: str = "supplier_response",
        rfp_template_schema: Optional[Dict] = None,
    ) -> Dict[str, Any]:
        """
        Full pipeline:
          1. Parse workbook → sheets
          2. Classify sheets → find pricing sheet
          3. Map columns → canonical schema
          4. Validate → confidence tier + flags
          5. Return structured result

        Args:
            file_bytes:           Raw bytes of .xlsx file
            supplier_name:        Supplier identifier
            source_type:          "rfp_template" | "supplier_response"
            rfp_template_schema:  If provided, use its column_map as reference

        Returns: dict with schema + validation + confidence_tier
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            logger.error("Failed to open workbook for %s: %s", supplier_name, e)
            return {"error": str(e), "supplier": supplier_name}

        # 1. Extract all sheets as row lists
        sheet_data: Dict[str, List[List[Any]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            sheet_data[sheet_name] = rows

        # 2. Classify all sheets
        classifications = classify_workbook_sheets(sheet_data)
        best_sheet      = get_best_pricing_sheet(classifications)

        if not best_sheet:
            logger.warning("No pricing sheet found in workbook for supplier: %s", supplier_name)
            return {
                "supplier":          supplier_name,
                "error":             "No pricing sheet detected",
                "sheet_classifications": {k: v["is_pricing_sheet"] for k, v in classifications.items()},
                "confidence_tier":   "LOW",
                "auto_ingest":       False,
            }

        # 3. Map to canonical schema
        cl          = classifications[best_sheet]
        rows        = sheet_data[best_sheet]
        col_map     = cl["column_map"]
        row_roles   = cl["row_roles"]

        # If RFP template provided, use its column_map to augment
        if rfp_template_schema and rfp_template_schema.get("raw_column_map"):
            for k, v in rfp_template_schema["raw_column_map"].items():
                if int(k) not in col_map:
                    col_map[int(k)] = v

        canonical = map_rows_to_schema(
            rows=rows,
            column_map=col_map,
            row_roles=row_roles,
            sheet_name=best_sheet,
            source_type=source_type,
        )

        # 4. Validate
        validation = validate_pricing_schema(canonical, supplier_name=supplier_name)

        return {
            "supplier":            supplier_name,
            "source_sheet":        best_sheet,
            "schema":              canonical,
            "validation":          validation,
            "confidence_tier":     validation["confidence_tier"],
            "auto_ingest":         validation["auto_ingest"],
            "review_needed":       validation["review_needed"],
            "sheet_classifications": {k: {
                "is_pricing": v["is_pricing_sheet"],
                "confidence": v["confidence"],
            } for k, v in classifications.items()},
        }

    # ── Ingest from text (existing behaviour) ────────────────────────────────

    def _parse_normalize(self, raw_pricing_data: list) -> dict:
        normalized = []
        for item in raw_pricing_data:
            try:
                parsed = parse_pricing_response(item.get("file_text", ""))
                for line in parsed.get("line_items", []):
                    uom        = line.get("unit", "each")
                    unit_price = line.get("unit_price") or line.get("price", 0)
                    line["normalized_unit_price"] = self._convert_uom(float(unit_price), uom)
                    line["base_uom"] = "each"
                normalized.append({"supplier": item.get("supplier"), "data": parsed})
            except Exception as e:
                logger.warning("parse_normalize failed for %s: %s", item.get("supplier"), e)
        return {"normalized_pricing": normalized}

    # ── TCO ──────────────────────────────────────────────────────────────────

    def _calculate_tco(
        self,
        unit_price: float,
        quantity: int,
        freight_pct: float = 0.0,
        duty_pct: float = 0.0,
        tooling_cost: float = 0.0,
    ) -> dict:
        base    = unit_price * quantity
        freight = base * freight_pct / 100
        duty    = base * duty_pct    / 100
        tco     = base + freight + duty + tooling_cost
        return {
            "base_cost":    round(base, 2),
            "freight":      round(freight, 2),
            "duty":         round(duty, 2),
            "tooling":      round(tooling_cost, 2),
            "tco":          round(tco, 2),
            "tco_per_unit": round(tco / quantity, 6) if quantity else 0,
        }

    # ── FX normalization ─────────────────────────────────────────────────────

    def _normalize_currency(self, prices: list, currencies: list = None) -> list:
        try:
            import requests
            rates = requests.get(
                f"https://api.exchangerate.host/latest?base={self.base_currency}",
                timeout=5,
            ).json().get("rates", {})
        except Exception:
            rates = {}
        for item in prices:
            src  = item.get("currency", self.base_currency)
            rate = rates.get(src, 1.0)
            item["price_normalized"]  = round(item.get("price", 0) / rate, 4)
            item["base_currency"]     = self.base_currency
        return prices

    # ── Validity ─────────────────────────────────────────────────────────────

    def _check_validity(self, quotes: list) -> list:
        today = datetime.utcnow().date()
        for q in quotes:
            expiry = q.get("valid_until")
            if expiry:
                try:
                    exp_date = date.fromisoformat(expiry)
                    delta    = (exp_date - today).days
                    q["validity_status"] = (
                        "expired"      if delta < 0  else
                        "expiring_soon" if delta <= 30 else
                        "valid"
                    )
                    q["days_to_expiry"] = delta
                except ValueError:
                    q["validity_status"] = "unknown"
        return quotes

    # ── UoM helper ───────────────────────────────────────────────────────────

    def _convert_uom(self, price: float, from_uom: str, to_uom: str = "each") -> float:
        from_f = UOM_FACTORS.get(from_uom.lower().strip(), 1.0)
        to_f   = UOM_FACTORS.get(to_uom.lower().strip(), 1.0)
        return round(price * from_f / to_f, 6)

    # ── Orchestrator entry point ──────────────────────────────────────────────

    def run(self, input: dict, context: dict = None) -> dict:
        action = input.get("action", "analyze")

        if action == "ingest_workbook":
            return self._ingest_workbook(
                file_bytes    = input["file_bytes"],
                supplier_name = input.get("supplier_name", ""),
                source_type   = input.get("source_type", "supplier_response"),
                rfp_template_schema = input.get("rfp_template_schema"),
            )

        if action == "tco":
            return self._calculate_tco(**{k: input[k] for k in input if k != "action"})

        if action == "validity":
            return {"quotes": self._check_validity(input.get("quotes", []))}

        if action == "currency":
            return {"prices": self._normalize_currency(input.get("prices", []))}

        # Default: full analyze flow (existing behavior)
        normalized  = self._parse_normalize(input.get("raw_pricing_data", []))
        suppliers   = []
        for entry in normalized["normalized_pricing"]:
            d = entry.get("data", {})
            suppliers.append({
                "supplier_name": entry["supplier"],
                "all_line_items": d.get("line_items", d.get("all_line_items", [])),
                "total_cost":    d.get("total_cost", 0),
            })
        cost_model = build_cost_model(suppliers)
        analysis   = analyze_pricing(cost_model)
        return {"cost_model": cost_model, "analysis": analysis, "normalized": normalized}
