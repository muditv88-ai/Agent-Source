#!/usr/bin/env python
"""Debug script to diagnose parser issues with Book2.xlsx"""

import sys
sys.path.insert(0, '.')

import openpyxl
from app.services.technical_parser import (
    detect_format, detect_sections, map_columns
)

# Replace this with actual path to Book2.xlsx
FILE_PATH = "uploads/projects/8427ef54-8533-4cd3-be99-9fa2c4ad1bfa/supplier_responses/Book2.xlsx"

try:
    wb = openpyxl.load_workbook(FILE_PATH, data_only=True)
    print("="*70)
    print("WORKBOOK DIAGNOSTIC FOR: {}".format(FILE_PATH))
    print("="*70)

    print("\nSheets found: {}".format(wb.sheetnames))

    fmt = detect_format(wb)
    print("Detected Format: {}".format(fmt))

    for ws in wb.worksheets:
        print("\n" + "="*70)
        print("SHEET: {}".format(ws.title))
        print("="*70)

        # Show structure
        print("\nFirst 15 rows of data:")
        for row_idx in range(1, min(16, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, 9):
                val = ws.cell(row_idx, col_idx).value
                if val is None:
                    row_values.append("")
                else:
                    # Safely encode to avoid unicode issues
                    safe_val = str(val)[:20].encode('ascii', 'replace').decode('ascii')
                    row_values.append(safe_val)
            try:
                print("  Row {:2d}: {}".format(row_idx, row_values))
            except UnicodeEncodeError:
                print("  Row {:2d}: [data with special chars]".format(row_idx))

        # Test section detection
        print("\nSection Detection:")
        sections = detect_sections(ws)
        if sections:
            print("  Found {} sections".format(len(sections)))
            for sec_idx, sec in enumerate(sections):
                print("\n  Section {}: '{}'".format(sec_idx+1, sec["section_name"]))
                print("    Header row: {}".format(sec["header_row_idx"]))
                print("    Data rows: {} to {}".format(sec["data_start_idx"], sec["data_end_idx"]))

                # Get header row
                header_row = [ws.cell(sec["header_row_idx"], col_idx + 1).value
                              for col_idx in range(ws.max_column)]
                print("    Headers: {}".format(header_row[:10]))

                # Map columns
                col_map = map_columns(header_row)
                print("    Mapped columns: {}".format(col_map))

                # Show first data row
                if sec["data_start_idx"] <= ws.max_row:
                    first_data = [ws.cell(sec["data_start_idx"], col_idx + 1).value
                                  for col_idx in range(min(8, ws.max_column))]
                    print("    First data row: {}".format(first_data))
        else:
            print("  No sections detected!")

except FileNotFoundError:
    print("ERROR: File not found: {}".format(FILE_PATH))
    print("\nUsage: Place Book2.xlsx in the uploads/ folder or update FILE_PATH")
except Exception as e:
    print("ERROR: {}".format(e))
    import traceback
    traceback.print_exc()
